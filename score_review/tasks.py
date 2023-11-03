from __future__ import absolute_import, unicode_literals
from celery import shared_task
from streams.scrapers.utilities import import_bot
from datetime import timedelta
from streams.models import Proxy
from api.utils import load_data_from_dataframe
from openpyxl import load_workbook
from django.utils import timezone
from api.models import TourURL, STREAM_CHOICES, Country, City, State
from datetime import datetime
import pytz
from api.models import TourOperator, Tour, Review
from django.db.models import Avg, Count


@shared_task(name="scrape")
def scrape(tour_id, *args, **kwargs):
    print('tour_id ', tour_id)
    tour = TourURL.objects.get(id=tour_id)
    stream_name = dict(STREAM_CHOICES)[tour.stream]
    Bot = import_bot(stream_name)
    Bot(tour)
    # return Bot(tour)  # TODO Uncomment this for debugging


@shared_task(name="scrape_delayed")
def scrape_delayed(tour_id, *args, **kwargs):
    eta = datetime.now() + timedelta(seconds=300)
    scrape.apply_async(args=[tour_id], eta=eta)

# Cleans orphan proxies
@shared_task(priority=0)
def clean_proxy():
    print('Clean proxy')
    proxies = Proxy.objects.all()
    for p in proxies:
        time_diff = timezone.now() - p.started_using
        if time_diff > timedelta(minutes=10):
            p.in_use = False
            p.save()


@shared_task(priority=0)
def create_city_country_state_relationship():
    tours = Tour.objects.all()

    for tour in tours:
        # Clean up strings
        country_name = tour.country.strip() if tour.country else None
        state_name = tour.state.strip() if tour.state else None
        city_name = tour.city.strip() if tour.city else None
        state = None

        # Country creation
        country, created = Country.objects.get_or_create(name=country_name)

        # Connect Country to Tour
        tour.country_id = country
        tour.save()

        if state_name:
            # State creation
            state, created = State.objects.get_or_create(name=state_name, country=country)

            # Connect State to Tour
            tour.state_id = state
            tour.save()

        # City creation
        city, created = City.objects.get_or_create(name=city_name, state=state)

        # Connect City to Tour
        tour.city_id = city
        tour.save()


@shared_task(priority=0)
def count_and_rating_task():
    tours = Tour.objects.all()
    for t in tours:
        reviews_data = t.reviews.aggregate(avg_rating=Avg('rating'), total_reviews=Count('id'))
        t.rating = reviews_data['avg_rating']
        t.n_reviews = reviews_data['total_reviews']
        t.save()


@shared_task(name="process_file_upload")
def process_file_upload(file_path):
    print('running process')

    # WE NEED '/home/app/' for the docker, don't delete it
    # workbook = load_workbook(filename= file_path, data_only=True)
    workbook = load_workbook(filename='/home/app/' + file_path, data_only=True)

    dict_streams = [
        {'stream_id': TourURL.TRIPADVISOR, 'stream': 'tripadvisor', 'row_ids': [7]},
        {'stream_id': TourURL.VIATOR, 'stream': 'viator', 'row_ids': [10, 11, 12, 13, 14, 15, 16, 17, 18, 19]},
        {'stream_id': TourURL.GETYOURGUIDE, 'stream': 'getyourguide',
         'row_ids': [20, 21, 22, 23, 24, 25, 26, 27, 28, 29]},
        {'stream_id': TourURL.TIQUETS, 'stream': 'tiqets', 'row_ids': [30, 31, 32, 33, 34, 35, 36, 37, 38, 39]},
        {'stream_id': TourURL.KLOOK, 'stream': 'klook', 'row_ids': [40, 41, 42, 43, 44, 45, 46, 47, 48]},
        {'stream_id': TourURL.EXPEDIA, 'stream': 'expedia', 'row_ids': [49, 50, 51, 52, 53, 54, 55, 56, 57, 58]},
        {'stream_id': TourURL.EXPEDIA, 'stream': 'google', 'row_ids': [59]},
    ]
    print('workbook loaded')
    sheet = workbook.active
    data = []
    rows_to_check = [0, 1, 4, 5, 6, 8]  # Checks if the important rows are populated
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        jump_loop = False

        if row[1] is None:
            continue

        for x in rows_to_check:
            if row[x] is None:
                jump_loop = True
        if jump_loop:
            continue

        date_entered = row[0]
        date_str = date_entered.strftime("%Y-%m-%d %H:%M:%S.%f")
        date_entered = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S.%f")
        date_entered = pytz.timezone('Europe/Rome').localize(date_entered)

        operator_data = {
            'operator': row[1],
            # 'category': row[2] if row[2] else None,
            'rating': row[3] if row[3] else None,
            'country': row[4],
            'state': row[5] if row[5] else None,
            'city': row[6],
            'email': row[9],
            'website': row[8],
            'date_created': date_entered,
            'stream_data': []
        }

        for d in dict_streams:
            for computed_value_id in d['row_ids']:
                computed_value = row[computed_value_id]
                if computed_value:
                    operator_data['stream_data'].append({
                        'stream': d['stream_id'],
                        'url': computed_value
                    })
        data.append(operator_data)

    load_data_from_dataframe(data)
